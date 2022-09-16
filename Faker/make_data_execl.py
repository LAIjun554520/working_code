# from openpyxl import workbook
# from faker import Faker
# import time
# now=time.time()
#
# wb = workbook.Workbook()
# faker = Faker("en_US")
# ws = wb.active
# ws.append(['id','age', 'name', 'am_pm', 'boolean', 'international_no', 'international_name', 'city', 'post_no', 'date'])
# for i in range(500):#注意：1.每次执行后都会覆盖之前的数据
#     ws.append([i+500, faker.random_int(),faker.name(),faker.am_pm(), faker.boolean(), faker.country_code(), faker.country(), faker.country(), faker.postcode(), faker.date_object()])
# #wb.save('c:/Users/Desktop/data.xlsx')#自定义文件存放位置，若无会则自动创建
# wb.save("execl_file"+str(int(round(now * 1000)))+".xlsx") # 默认存放在当前project下，若无会自动创建
#

from openpyxl import workbook
from faker import Faker
import time
now=time.time()

wb = workbook.Workbook()
faker = Faker("zh-CN")
ws = wb.active
ws.append(['序号', '姓名', '工作', '公司', '市', '区', '街道名', '公司性质', '词语'])
for i in range(500):#注意：1.每次执行后都会覆盖之前的数据
    ws.append([i+1500, faker.name(),faker.job(), faker.company(), faker.city_suffix(), faker.district(), faker.street_name(), faker.company_suffix(), faker.word()])
#wb.save('c:/Users/Desktop/data.xlsx')#自定义文件存放位置，若无会则自动创建
wb.save("execl_file"+str(int(round(now * 1000)))+".xlsx") # 默认存放在当前project下，若无会自动创建

#
# from openpyxl import workbook
# from faker import Faker
#
# wb = workbook.Workbook()
# fake = Faker("zh-CN")
# ws = wb.active
# ws.append(['id','name','ATINYINT','ASMALLINT','AINT','ABIGINT','AFLOAT','ADOUBLE','ADECIMAL','ABOOLEAN','ASTRING','ACHAR','AVARCHAR','AVARCHAR2','ADATE','ATIMESTAME','ATIME'])
# for i in range(500):#注意：1.每次执行后都会覆盖之前的数据
#     ws.append([i,
#              fake.name(),
#              fake.pyfloat(left_digits=2,right_digits=2,positive=True),
#              fake.pyfloat(left_digits=2,right_digits=3,positive=True),
#              fake.random_digit(),
#              fake.random_int(),
#              fake.pyfloat(left_digits=2, right_digits=4, positive=True),
#              fake.random_int(),
#              fake.pyfloat(left_digits=2, right_digits=4, positive=True),
#              fake.boolean(),
#              fake.district(),
#              fake.street_name(),
#              fake.company_suffix(),
#              fake.sentence(),
#              fake.past_date(),
#              fake.future_datetime(),
#              fake.time()])
# #wb.save('c:/Users/Desktop/data.xlsx')#自定义文件存放位置，若无会则自动创建
# wb.save('/Users/simons/PycharmProjects/TEST/studio_performance_test/Faker/execl_type.xlsx') # 默认存放在当前project下，若无会自动创建